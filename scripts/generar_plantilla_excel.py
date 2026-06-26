"""
Script para generar la plantilla XLSX de carga masiva de productos MIKITECH.
Ejecutar: python scripts/generar_plantilla_excel.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ─── Colores MIKITECH ─────────────────────────────────────────────────────────
AZUL_PRIMARIO = '1D4ED8'
NAVY_DARK     = '0F172A'
GRIS_BORDE    = 'CBD5E1'
VERDE_OK      = 'DCFCE7'
BLANCO        = 'FFFFFF'

# ─── Borde fino compartido ───────────────────────────────────────────────────
def borde_fino():
    lado = Side(style='thin', color=GRIS_BORDE)
    return Border(left=lado, right=lado, top=lado, bottom=lado)

# ─── Definición de columnas ──────────────────────────────────────────────────
COLUMNAS = [
    ('nombre',               'REQUERIDO', 'Nombre del producto. Ej: Mouse Gamer Logitech G502'),
    ('categoria_slug',       'REQUERIDO', 'Slug de la categoría. Ej: mouse, teclados, monitores'),
    ('precio',               'REQUERIDO', 'Precio en pesos. Solo números. Ej: 150000'),
    ('existencias',          'REQUERIDO', 'Cantidad en stock. Solo enteros. Ej: 25'),
    ('marca',                'OPCIONAL',  'Marca del producto. Ej: Logitech'),
    ('modelo',               'OPCIONAL',  'Modelo específico. Ej: G502 HERO'),
    ('sku',                  'OPCIONAL',  'Código SKU único. Ej: LOG-G502-001'),
    ('descripcion_corta',    'OPCIONAL',  'Resumen en máx 500 caracteres'),
    ('descripcion',          'OPCIONAL',  'Descripción completa del producto'),
    ('url_imagen',           'OPCIONAL',  'URL directa a la imagen principal'),
    ('es_destacado',         'OPCIONAL',  'Escribe: si o no. Default: no'),
    ('descuento_porcentaje', 'OPCIONAL',  'Número 0–100. Default: 0'),
]

# ─── Filas de ejemplo ────────────────────────────────────────────────────────
EJEMPLOS = [
    [
        'Mouse Gamer Logitech G502',
        'mouse',
        '189900',
        '15',
        'Logitech',
        'G502 HERO',
        'LOG-G502-001',
        'Mouse gaming de alta precisión con 11 botones programables',
        'El Logitech G502 HERO es el mouse definitivo para gamers. Sensor HERO 25K.',
        'https://resource.logitechg.com/w_692/content/dam/gaming/g502-x-gallery-1.png',
        'si',
        '10',
    ],
    [
        'Teclado Mecánico Redragon K552',
        'teclados',
        '145000',
        '8',
        'Redragon',
        'K552 KUMARA',
        'RED-K552-001',
        'Teclado mecánico TKL RGB con switches rojos',
        'Teclado mecánico compacto TKL con switches Red lineales y retroiluminación RGB.',
        '',
        'no',
        '0',
    ],
    [
        'Monitor Gaming Samsung 27" QHD',
        'monitores',
        '890000',
        '3',
        'Samsung',
        'Odyssey G5',
        'SAM-OG5-27',
        'Monitor curvo 1440p 165Hz para gaming competitivo',
        'Panel VA curvo 1000R, QHD 2560x1440, 165Hz, AMD FreeSync Premium, 1ms GTG.',
        '',
        'si',
        '5',
    ],
]


def aplicar_celda(celda, valor=None, bold=False, size=10, color_texto=NAVY_DARK,
                  color_fondo=BLANCO, alineacion='left', wrap=False, borde=True):
    if valor is not None:
        celda.value = valor
    celda.font = Font(name='Calibri', bold=bold, size=size, color=color_texto)
    celda.fill = PatternFill('solid', fgColor=color_fondo)
    celda.alignment = Alignment(horizontal=alineacion, vertical='center', wrap_text=wrap)
    if borde:
        celda.border = borde_fino()


def crear_hoja_productos(wb):
    ws = wb.active
    ws.title = 'Productos'

    # ── Fila 1: Título ────────────────────────────────────────────────────────
    ws.merge_cells('A1:L1')
    aplicar_celda(ws['A1'],
                  valor='MIKITECH — Plantilla de Carga Masiva de Productos',
                  bold=True, size=14,
                  color_texto=BLANCO, color_fondo=NAVY_DARK,
                  alineacion='center', borde=False)
    ws.row_dimensions[1].height = 34

    # ── Fila 2: Instrucción rápida ────────────────────────────────────────────
    ws.merge_cells('A2:L2')
    aplicar_celda(ws['A2'],
                  valor='Completa los campos REQUERIDOS (azul). Los OPCIONALES (gris) pueden dejarse vacíos. NO modifiques los encabezados.',
                  size=10, color_texto='374151', color_fondo='E0F2FE',
                  alineacion='left', wrap=True, borde=False)
    ws.row_dimensions[2].height = 22

    # ── Filas 3-5: Cabeceras ──────────────────────────────────────────────────
    for idx, (campo, tipo, desc) in enumerate(COLUMNAS, start=1):
        col = get_column_letter(idx)
        es_req = (tipo == 'REQUERIDO')

        # Fila 3 — Nombre del campo
        aplicar_celda(ws[f'{col}3'],
                      valor=campo, bold=True, size=11,
                      color_texto=BLANCO,
                      color_fondo=AZUL_PRIMARIO if es_req else '64748B',
                      alineacion='center')

        # Fila 4 — Tipo
        aplicar_celda(ws[f'{col}4'],
                      valor=tipo, bold=True, size=9,
                      color_texto='1D4ED8' if es_req else '6B7280',
                      color_fondo='EFF6FF' if es_req else 'F9FAFB',
                      alineacion='center')

        # Fila 5 — Descripción
        aplicar_celda(ws[f'{col}5'],
                      valor=desc, size=8,
                      color_texto='475569', color_fondo='F8FAFC',
                      alineacion='left', wrap=True)

    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 42

    # ── Filas 6-8: Ejemplos ───────────────────────────────────────────────────
    colores_ejemplo = [VERDE_OK, 'F0FDF4', 'DCFCE7']
    for fila_idx, (ejemplo, color) in enumerate(zip(EJEMPLOS, colores_ejemplo), start=6):
        for col_idx, valor in enumerate(ejemplo, start=1):
            aplicar_celda(ws[f'{get_column_letter(col_idx)}{fila_idx}'],
                          valor=valor, size=10,
                          color_texto='1E293B', color_fondo=color,
                          alineacion='left', wrap=True)
        ws.row_dimensions[fila_idx].height = 20

    # Separador visual entre ejemplos y datos
    ws.row_dimensions[9].height = 4

    # ── Filas 10-60: Datos vacíos ─────────────────────────────────────────────
    for fila_idx in range(10, 61):
        color_fila = BLANCO if fila_idx % 2 == 0 else 'F8FAFC'
        for col_idx in range(1, len(COLUMNAS) + 1):
            aplicar_celda(ws[f'{get_column_letter(col_idx)}{fila_idx}'],
                          size=10, color_fondo=color_fila, alineacion='left')
        ws.row_dimensions[fila_idx].height = 18

    # ── Anchos de columna ─────────────────────────────────────────────────────
    anchos = [38, 20, 14, 13, 20, 20, 22, 42, 52, 48, 13, 22]
    for idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho

    # Congelar paneles a partir de la fila 6
    ws.freeze_panes = 'A6'
    return ws


def crear_hoja_categorias(wb):
    ws = wb.create_sheet('Categorías')

    ws.merge_cells('A1:B1')
    aplicar_celda(ws['A1'],
                  valor='Categorías disponibles (usa el slug en la columna categoria_slug)',
                  bold=True, size=12,
                  color_texto=BLANCO, color_fondo=NAVY_DARK,
                  alineacion='center', borde=False)
    ws.row_dimensions[1].height = 30

    for col, titulo in [('A', 'Nombre de Categoría'), ('B', 'Slug a usar en el archivo')]:
        aplicar_celda(ws[f'{col}2'],
                      valor=titulo, bold=True, size=10,
                      color_texto=BLANCO, color_fondo=AZUL_PRIMARIO,
                      alineacion='center')

    ws.merge_cells('A3:B3')
    aplicar_celda(ws['A3'],
                  valor='Los slugs se generan automáticamente al crear categorías en el admin. Copia el slug exacto.',
                  size=9, color_texto='475569', color_fondo='E0F2FE',
                  alineacion='left', wrap=True, borde=False)
    ws.row_dimensions[3].height = 26

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 30
    return ws


def crear_hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones')

    filas = [
        ('INSTRUCCIONES DE USO — MIKITECH Carga Masiva', True, 14, BLANCO, NAVY_DARK),
        ('', False, 10, '000000', BLANCO),
        ('1. PREPARACIÓN DEL ARCHIVO', True, 11, '1D4ED8', 'EFF6FF'),
        ('   • Trabaja solo en la hoja "Productos". No toques las otras hojas.', False, 10, '374151', BLANCO),
        ('   • Los campos AZULES son OBLIGATORIOS. Los GRISES son opcionales.', False, 10, '374151', BLANCO),
        ('   • No elimines ni renombres las filas 3, 4 o 5 (encabezados).', False, 10, '374151', BLANCO),
        ('   • Las filas 6, 7 y 8 son ejemplos; puedes borrarlas antes de subir.', False, 10, '374151', BLANCO),
        ('', False, 10, '000000', BLANCO),
        ('2. REGLAS POR CAMPO', True, 11, '1D4ED8', 'EFF6FF'),
        ('   • nombre: Obligatorio. Máximo 200 caracteres.', False, 10, '374151', BLANCO),
        ('   • categoria_slug: Obligatorio. Debe existir en el admin. Ver hoja "Categorías".', False, 10, '374151', BLANCO),
        ('   • precio: Obligatorio. Solo números enteros sin puntos ni comas. Ej: 150000', False, 10, 'DC2626', 'FEF2F2'),
        ('   • existencias: Obligatorio. Solo números enteros. Ej: 25', False, 10, '374151', BLANCO),
        ('   • es_destacado: Escribe exactamente: si o no (minúscula).', False, 10, '374151', BLANCO),
        ('   • descuento_porcentaje: Número entre 0 y 100. Deja vacío para 0.', False, 10, '374151', BLANCO),
        ('   • sku: Si lo escribes, debe ser único. Si está vacío, se omite.', False, 10, '374151', BLANCO),
        ('', False, 10, '000000', BLANCO),
        ('3. PROCESO DE CARGA', True, 11, '1D4ED8', 'EFF6FF'),
        ('   • Inicia sesión en el Panel Admin → Productos → Carga Masiva.', False, 10, '374151', BLANCO),
        ('   • Arrastra el archivo .xlsx o usa el selector para subirlo.', False, 10, '374151', BLANCO),
        ('   • El sistema procesará cada fila y mostrará un reporte al finalizar.', False, 10, '374151', BLANCO),
        ('   • Las filas con error indicarán el motivo exacto del problema.', False, 10, '374151', BLANCO),
        ('', False, 10, '000000', BLANCO),
        ('4. ERRORES COMUNES', True, 11, '1D4ED8', 'EFF6FF'),
        ('   ❌  categoria_slug no encontrado → Verifica la hoja "Categorías"', False, 10, 'DC2626', 'FEF2F2'),
        ('   ❌  precio inválido → Usa solo números. Sin puntos ($), comas ni espacios.', False, 10, 'DC2626', 'FEF2F2'),
        ('   ❌  SKU duplicado → Ese código ya existe en otro producto.', False, 10, 'DC2626', 'FEF2F2'),
        ('   ❌  nombre vacío → El nombre es un campo obligatorio.', False, 10, 'DC2626', 'FEF2F2'),
    ]

    for i, (texto, bold, size, c_text, c_bg) in enumerate(filas, start=1):
        celda = ws[f'A{i}']
        aplicar_celda(celda, valor=texto, bold=bold, size=size,
                      color_texto=c_text, color_fondo=c_bg,
                      alineacion='left', wrap=True, borde=False)
        ws.row_dimensions[i].height = 22 if texto else 8

    ws.column_dimensions['A'].width = 85
    return ws


def main():
    wb = openpyxl.Workbook()
    crear_hoja_productos(wb)
    crear_hoja_categorias(wb)
    crear_hoja_instrucciones(wb)

    ruta_salida = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        'servidor-y-logica',
        'static',
        'plantilla_carga_masiva_mikitech.xlsx'
    )
    wb.save(ruta_salida)
    print(f"[OK] Plantilla generada en: {ruta_salida}")
    return ruta_salida


if __name__ == '__main__':
    main()
