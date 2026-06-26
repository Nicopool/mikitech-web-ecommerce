"""Vistas del panel administrativo con gateway de seguridad — MIKITECH"""

import uuid
from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.views.decorators.http import require_POST
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.contrib import messages
from products.models import Producto, Categoria, ImagenProducto
from interactions.models import Reseña, Pedido
from users.models import Perfil, Notificacion, InvitacionAdmin, Rol
from users.decorators import requiere_permiso
import os
from django.core.files.storage import default_storage


from functools import wraps

def guardar_archivo_hibrido(nombre_unico, archivo):
    """
    Sube el archivo a Supabase Storage. Si falla, lo guarda localmente en el FileSystemStorage.
    """
    from users.supabase_auth import subir_a_supabase_storage
    try:
        datos = archivo.read()
        archivo.seek(0)
        url_publica, error = subir_a_supabase_storage(nombre_unico, datos, archivo.content_type)
        if url_publica:
            return url_publica
        else:
            print(f"[!] Error Supabase Storage: {error}, usando fallback local...")
    except Exception as ex:
        print(f"[!] Excepción Supabase Storage: {ex}, usando fallback local...")
        
    ruta_guardada = default_storage.save(nombre_unico, archivo)
    return f"{settings.MEDIA_URL}{ruta_guardada}"

CODIGO_ADMIN = getattr(settings, 'ADMIN_GATEWAY_CODE', '')


def requerir_administrador(función_vista):
    """
    Decorador para rutas administrativas.

    Verifica en orden:
    1. Que exista una sesión activa con usuario_id.
    2. Que el rol de sesión sea 'admin'.
    """
    @wraps(función_vista)
    def envoltura(petición, *args, **kwargs):
        # Cerrojo 1: debe haber sesión activa
        if not petición.session.get('usuario_id'):
            return redirect('/admin-panel/login/')

        # Cerrojo 2: el rol de sesión (sincronizado con BD) debe ser 'admin'
        if petición.session.get('rol_usuario') != 'admin':
            return redirect('/admin-panel/login/')

        return función_vista(petición, *args, **kwargs)
    return envoltura


def pasarela(petición):
    """Redirecciona al login de administración al estar descontinuada la pasarela."""
    return redirect('/admin-panel/login/')


def login_administrador(petición):
    """Inicio de sesión exclusivo para administradores."""
    if petición.session.get('rol_usuario') == 'admin':
        return redirect('/admin-panel/')

    if petición.method == 'POST':
        from users.supabase_auth import iniciar_sesion_usuario
        correo = petición.POST.get('correo', '').strip()
        clave = petición.POST.get('clave', '')
        terminos = petición.POST.get('terminos')

        if not terminos:
            return render(petición, 'admin_panel/login.html', {
                'error': 'Debes aceptar los Términos y Condiciones.',
                'correo': correo,
                'titulo_pagina': 'Login Administrador — MIKITECH',
            })

        datos, error = iniciar_sesion_usuario(correo, clave)

        if error:
            return render(petición, 'admin_panel/login.html', {
                'error': f'Credenciales incorrectas: {error}',
                'correo': correo,
                'titulo_pagina': 'Login Administrador — MIKITECH',
            })

        id_usuario = datos.get('user', {}).get('id')
        try:
            perfil = Perfil.objects.get(id=id_usuario)
            if perfil.rol != 'admin':
                return render(petición, 'admin_panel/login.html', {
                    'error': 'Acceso denegado: Este usuario no tiene permisos de administrador.',
                    'correo': correo,
                    'titulo_pagina': 'Login Administrador — MIKITECH',
                })
            petición.session['usuario_id'] = id_usuario
            petición.session['token_acceso'] = datos.get('access_token')
            petición.session['rol_usuario'] = 'admin'
            petición.session['pasarela_administrador_superada'] = True  # Confirmar autorización completa
            petición.session['nombre_usuario'] = perfil.nombre_usuario
            petición.session['avatar_url'] = perfil.url_avatar or ''
            petición.session.modified = True
            messages.success(petición, f'✅ ¡Haz ingresado exitosamente! Bienvenido al Panel de Administración, {perfil.nombre_usuario}.')
            return redirect('/admin-panel/')
        except Perfil.DoesNotExist:
            user_data = datos.get('user', {})
            metadata = user_data.get('user_metadata', {})
            role = metadata.get('role')
            
            if role == 'admin':
                nombre_completo = metadata.get('full_name', 'Administrador')
                nombre_usuario = metadata.get('username', correo.split('@')[0])
                perfil = Perfil.objects.create(
                    id=id_usuario,
                    nombre_completo=nombre_completo,
                    nombre_usuario=nombre_usuario,
                    rol='admin',
                    esta_activo=True
                )
                petición.session['usuario_id'] = id_usuario
                petición.session['token_acceso'] = datos.get('access_token')
                petición.session['rol_usuario'] = 'admin'
                petición.session['pasarela_administrador_superada'] = True  # Confirmar autorización completa
                petición.session['nombre_usuario'] = perfil.nombre_usuario
                petición.session['avatar_url'] = perfil.url_avatar or ''
                petición.session.modified = True
                messages.success(petición, f'✅ ¡Haz ingresado exitosamente! Bienvenido al Panel de Administración, {perfil.nombre_usuario}.')
                return redirect('/admin-panel/')
            else:
                return render(petición, 'admin_panel/login.html', {
                    'error': 'Acceso denegado: Tu usuario no tiene el rol de administrador asignado.',
                    'correo': correo,
                    'titulo_pagina': 'Login Administrador — MIKITECH',
                })

    return render(petición, 'admin_panel/login.html', {
        'titulo_pagina': 'Login Administrador — MIKITECH',
    })


def cerrar_sesion_administrador(petición):
    """Cierra la sesión del administrador."""
    petición.session.flush()
    return redirect('/admin-panel/login/')


@requerir_administrador
@requiere_permiso('ver_reportes')
def tablero_administrador(petición):
    """Estadísticas principales del panel con filtro de tiempo (Semana, Mes, Año, Todo)."""
    from django.db.models import Sum, F
    from django.utils import timezone
    from datetime import timedelta
    
    filtro_tiempo = petición.GET.get('filtro_tiempo', 'mes')
    if filtro_tiempo not in ['semana', 'mes', 'anio', 'todo']:
        filtro_tiempo = 'mes'
        
    ahora = timezone.now()
    fecha_inicio = None
    
    if filtro_tiempo == 'semana':
        fecha_inicio = ahora - timedelta(days=7)
    elif filtro_tiempo == 'mes':
        fecha_inicio = ahora - timedelta(days=30)
    elif filtro_tiempo == 'anio':
        fecha_inicio = ahora - timedelta(days=365)
        
    # Calcular ingresos estimados del periodo seleccionado
    pedido_qs = Pedido.objects.filter(estado__in=['delivered', 'shipped'])
    if fecha_inicio:
        pedido_qs = pedido_qs.filter(creado_el__gte=fecha_inicio)
        
    ingresos_estimados = pedido_qs.aggregate(total=Sum('monto_total'))['total'] or 0
    
    # Calcular ingresos del periodo anterior para porcentaje de cambio
    porcentaje_cambio = 0
    comparacion_texto = "vs periodo anterior"
    
    if fecha_inicio:
        duracion = ahora - fecha_inicio
        fecha_inicio_anterior = fecha_inicio - duracion
        
        ingresos_periodo_anterior = Pedido.objects.filter(
            estado__in=['delivered', 'shipped'],
            creado_el__range=[fecha_inicio_anterior, fecha_inicio]
        ).aggregate(total=Sum('monto_total'))['total'] or 0
        
        if ingresos_periodo_anterior > 0:
            porcentaje_cambio = ((ingresos_estimados - ingresos_periodo_anterior) / ingresos_periodo_anterior) * 100
        else:
            porcentaje_cambio = 100 if ingresos_estimados > 0 else 0
            
        if filtro_tiempo == 'semana':
            comparacion_texto = "vs semana anterior"
        elif filtro_tiempo == 'mes':
            comparacion_texto = "vs mes anterior"
        elif filtro_tiempo == 'anio':
            comparacion_texto = "vs año anterior"
    else:
        comparacion_texto = "en total histórico"
        porcentaje_cambio = 100
        
    # Calcular datos del gráfico
    chart_labels = []
    chart_data = []
    
    if filtro_tiempo == 'semana':
        for i in range(6, -1, -1):
            dia = ahora.date() - timedelta(days=i)
            label = dia.strftime('%d %b')
            chart_labels.append(label)
            
            ingresos_dia = Pedido.objects.filter(
                estado__in=['delivered', 'shipped'],
                creado_el__date=dia
            ).aggregate(total=Sum('monto_total'))['total'] or 0
            chart_data.append(float(ingresos_dia))
            
    elif filtro_tiempo == 'mes':
        for i in range(3, -1, -1):
            fin_periodo = ahora - timedelta(days=i*7)
            inicio_periodo = fin_periodo - timedelta(days=6)
            label = f"Semana {4-i}"
            chart_labels.append(label)
            
            ingresos_periodo = Pedido.objects.filter(
                estado__in=['delivered', 'shipped'],
                creado_el__range=[inicio_periodo, fin_periodo]
            ).aggregate(total=Sum('monto_total'))['total'] or 0
            chart_data.append(float(ingresos_periodo))
            
    else:  # 'anio' o 'todo'
        for i in range(11, -1, -1):
            año_ref = ahora.year
            mes_ref = ahora.month - i
            while mes_ref <= 0:
                mes_ref += 12
                año_ref -= 1
            
            import calendar
            _, ultimo_dia = calendar.monthrange(año_ref, mes_ref)
            
            inicio_mes = timezone.make_aware(timezone.datetime(año_ref, mes_ref, 1, 0, 0, 0))
            fin_mes = timezone.make_aware(timezone.datetime(año_ref, mes_ref, ultimo_dia, 23, 59, 59))
            
            nombre_mes = inicio_mes.strftime('%b %y')
            chart_labels.append(nombre_mes)
            
            ingresos_mes = Pedido.objects.filter(
                estado__in=['delivered', 'shipped'],
                creado_el__range=[inicio_mes, fin_mes]
            ).aggregate(total=Sum('monto_total'))['total'] or 0
            chart_data.append(float(ingresos_mes))
            
    valor_total_inv = Producto.objects.aggregate(total=Sum(F('precio') * F('existencias')))['total'] or 0
    
    estadísticas = {
        'total_productos': Producto.objects.count(),
        'productos_activos': Producto.objects.filter(esta_activo=True).count(),
        'total_categorias': Categoria.objects.count(),
        'total_usuarios': Perfil.objects.count(),
        'total_resenas': Reseña.objects.count(),
        'stock_bajo': Producto.objects.filter(existencias__lte=5, esta_activo=True).count(),
        'valor_total_inventario': valor_total_inv,
        'ingresos_estimados': ingresos_estimados,
    }
    productos_recientes = Producto.objects.order_by('-creado_el')[:5]
    reseñas_recientes = Reseña.objects.select_related('producto', 'usuario').order_by('-creado_el')[:5]

    return render(petición, 'admin_panel/dashboard.html', {
        'estadisticas': estadísticas,
        'productos_recientes': productos_recientes,
        'resenas_recientes': reseñas_recientes,
        'porcentaje_cambio': porcentaje_cambio,
        'comparacion_texto': comparacion_texto,
        'filtro_tiempo': filtro_tiempo,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'titulo_pagina': 'Tablero de Control — MIKITECH',
    })


@requerir_administrador
def gestion_productos(petición):
    """Lista de productos para gestión administrativa."""
    productos = Producto.objects.select_related('categoria').order_by('-creado_el')
    
    from django.db.models import Sum, F
    valor_total_inv = productos.aggregate(total=Sum(F('precio') * F('existencias')))['total'] or 0
    
    return render(petición, 'admin_panel/products.html', {
        'productos': productos,
        'valor_total_inventario': valor_total_inv,
        'titulo_pagina': 'Gestión de Productos — MIKITECH',
    })


@requerir_administrador
def crear_producto(petición):
    """Procesamiento y vista para la creación de nuevos productos."""
    categorías = Categoria.objects.all().order_by('nombre')
    if petición.method == 'POST':
        nombre = petición.POST.get('nombre', '').strip()
        id_categoria = petición.POST.get('id_categoria', '')
        precio = petición.POST.get('precio', '0')
        existencias = petición.POST.get('existencias', '0')
        marca = petición.POST.get('marca', '').strip()
        descripcion = petición.POST.get('descripcion', '').strip()
        descripcion_corta = petición.POST.get('descripcion_corta', '').strip()
        es_destacado = petición.POST.get('es_destacado') == 'on'
        descuento_porcentaje = int(petición.POST.get('descuento_porcentaje', '0') or '0')
        descuento_expira_str = petición.POST.get('descuento_expira_el', '').strip()
        descuento_expira_el = None
        if descuento_expira_str:
            from django.utils import timezone
            from datetime import datetime
            try:
                descuento_expira_el = datetime.fromisoformat(descuento_expira_str).astimezone(timezone.utc)
            except ValueError:
                pass

        if nombre:
            from django.utils.text import slugify
            base_enlace = slugify(nombre)
            enlace = base_enlace
            contador = 1
            while Producto.objects.filter(enlace=enlace).exists():
                enlace = f"{base_enlace}-{contador}"
                contador += 1

            try:
                cat = Categoria.objects.get(id=id_categoria)
                nuevo_prod = Producto.objects.create(
                    id=uuid.uuid4(),
                    categoria=cat,
                    nombre=nombre,
                    enlace=enlace,
                    precio=float(precio),
                    existencias=int(existencias),
                    marca=marca,
                    descripcion=descripcion,
                    descripcion_corta=descripcion_corta,
                    es_destacado=es_destacado,
                    descuento_porcentaje=descuento_porcentaje,
                    descuento_expira_el=descuento_expira_el,
                    esta_activo=True,
                )
                
                # Imagen Principal
                if 'archivo_imagen' in petición.FILES:
                    archivo = petición.FILES['archivo_imagen']
                    nombre_unico = f"products/{uuid.uuid4()}{os.path.splitext(archivo.name)[1]}"
                    nuevo_prod.url_imagen_principal = guardar_archivo_hibrido(nombre_unico, archivo)
                    nuevo_prod.save()

                # Galería de Imágenes
                if 'archivos_galeria' in petición.FILES:
                    for f in petición.FILES.getlist('archivos_galeria'):
                        nombre_extra = f"products/gallery/{uuid.uuid4()}{os.path.splitext(f.name)[1]}"
                        url_extra = guardar_archivo_hibrido(nombre_extra, f)
                        ImagenProducto.objects.create(
                            producto=nuevo_prod,
                            url_imagen=url_extra
                        )
                
                messages.success(petición, f'Producto "{nombre}" creado con éxito.')
                messages.success(petición, f'Producto "{nombre}" creado con éxito.')
                return redirect('/admin-panel/productos/')
            except Exception as e:
                return render(petición, 'admin_panel/product_form.html', {
                    'categorias': categorías,
                    'error': f'Error al guardar el producto: {str(e)}',
                    'titulo_pagina': 'Crear Producto — MIKITECH',
                })

    return render(petición, 'admin_panel/product_form.html', {
        'categorias': categorías,
        'titulo_pagina': 'Crear Producto — MIKITECH',
    })


@requerir_administrador
def editar_producto(petición, id_producto):
    """Formulario para la edición de productos existentes."""
    producto = get_object_or_404(Producto, id=id_producto)
    categorías = Categoria.objects.all().order_by('nombre')
    if petición.method == 'POST':
        producto.nombre = petición.POST.get('nombre', producto.nombre).strip()
        producto.precio = float(petición.POST.get('precio', producto.precio))
        producto.existencias = int(petición.POST.get('existencias', producto.existencias))
        producto.marca = petición.POST.get('marca', producto.marca)
        producto.descripcion = petición.POST.get('descripcion', producto.descripcion)
        producto.descripcion_corta = petición.POST.get('descripcion_corta', producto.descripcion_corta)
        
        # Imagen Principal Local
        if 'archivo_imagen' in petición.FILES:
            archivo = petición.FILES['archivo_imagen']
            nombre_unico = f"products/{uuid.uuid4()}{os.path.splitext(archivo.name)[1]}"
            producto.url_imagen_principal = guardar_archivo_hibrido(nombre_unico, archivo)
            
        # Galería: Nuevas Imágenes
        if 'archivos_galeria' in petición.FILES:
            for f in petición.FILES.getlist('archivos_galeria'):
                nombre_extra = f"products/gallery/{uuid.uuid4()}{os.path.splitext(f.name)[1]}"
                url_extra = guardar_archivo_hibrido(nombre_extra, f)
                ImagenProducto.objects.create(
                    producto=producto,
                    url_imagen=url_extra
                )

        # Galería: Eliminaciones
        ids_eliminar = petición.POST.getlist('eliminar_imagenes')
        if ids_eliminar:
            ImagenProducto.objects.filter(id__in=ids_eliminar, producto=producto).delete()

        producto.es_destacado = petición.POST.get('es_destacado') == 'on'
        # Al editar producto, el estado activo viene explícito
        producto.esta_activo = petición.POST.get('esta_activo') == 'on'
        # Descuento
        producto.descuento_porcentaje = int(petición.POST.get('descuento_porcentaje', '0') or '0')
        desc_exp_str = petición.POST.get('descuento_expira_el', '').strip()
        if desc_exp_str:
            from django.utils import timezone
            from datetime import datetime
            try:
                producto.descuento_expira_el = datetime.fromisoformat(desc_exp_str).astimezone(timezone.utc)
            except ValueError:
                producto.descuento_expira_el = None
        else:
            producto.descuento_expira_el = None
        id_cat = petición.POST.get('id_categoria')
        if id_cat:
            try:
                producto.categoria = Categoria.objects.get(id=id_cat)
            except Categoria.DoesNotExist:
                pass
        producto.save()
        messages.success(petición, f'Cambios en "{producto.nombre}" guardados.')
        return redirect('/admin-panel/productos/')

    return render(petición, 'admin_panel/product_form.html', {
        'producto': producto,
        'categorias': categorías,
        'titulo_pagina': f'Editar: {producto.nombre} — MIKITECH',
    })


@requerir_administrador
@require_POST
def eliminar_producto(petición, id_producto):
    """Proceso de eliminación física de un producto."""
    producto = get_object_or_404(Producto, id=id_producto)
    nombre = producto.nombre
    producto.delete()
    messages.success(petición, f'Producto "{nombre}" eliminado satisfactoriamente.')
    return redirect('/admin-panel/productos/')


@requerir_administrador
def gestion_categorias(petición):
    """Visualización y control de las categorías del catálogo."""
    categorías = Categoria.objects.all().order_by('nombre')
    return render(petición, 'admin_panel/categories.html', {
        'categorias': categorías,
        'titulo_pagina': 'Gestión de Categorías — MIKITECH',
    })


@requerir_administrador
def crear_categoria(petición):
    """Añadir nuevas agrupaciones de productos."""
    if petición.method == 'POST':
        from django.utils.text import slugify
        nombre = petición.POST.get('nombre', '').strip()
        descripcion = petición.POST.get('descripcion', '').strip()
        icono = petición.POST.get('icono', '').strip()
        url_imagen = petición.POST.get('url_imagen', '').strip()
        
        # Procesar Archivo Local si existe
        if 'archivo_imagen' in petición.FILES:
            archivo = petición.FILES['archivo_imagen']
            nombre_unico = f"categories/{uuid.uuid4()}{os.path.splitext(archivo.name)[1]}"
            url_imagen = guardar_archivo_hibrido(nombre_unico, archivo)
        
        if not nombre:
            return render(petición, 'admin_panel/category_form.html', {
                'error': 'El nombre de la categoría es obligatorio.',
                'titulo_pagina': 'Crear Categoría — MIKITECH'
            })

        base_enlace = slugify(nombre)
        enlace = base_enlace
        contador = 1
        while Categoria.objects.filter(enlace=enlace).exists():
            enlace = f"{base_enlace}-{contador}"
            contador += 1

        try:
            Categoria.objects.create(
                nombre=nombre, 
                enlace=enlace, 
                descripcion=descripcion, 
                icono=icono,
                url_imagen=url_imagen
            )
            messages.success(petición, f'Categoría "{nombre}" establecida.')
            return redirect('/admin-panel/categorias/')
        except Exception as e:
            return render(petición, 'admin_panel/category_form.html', {
                'error': f'Imposible crear categoría: {str(e)}',
                'titulo_pagina': 'Crear Categoría — MIKITECH'
            })

    return render(petición, 'admin_panel/category_form.html', {'titulo_pagina': 'Crear Categoría — MIKITECH'})


@requerir_administrador
def editar_categoria(petición, id_cat):
    """Modificar propiedades de una categoría existente."""
    categoría = get_object_or_404(Categoria, id=id_cat)
    if petición.method == 'POST':
        categoría.nombre = petición.POST.get('nombre', categoría.nombre).strip()
        categoría.descripcion = petición.POST.get('descripcion', categoría.descripcion).strip()
        categoría.icono = petición.POST.get('icono', categoría.icono).strip()
        
        # Procesar Archivo Local si existe
        if 'archivo_imagen' in petición.FILES:
            archivo = petición.FILES['archivo_imagen']
            nombre_unico = f"categories/{uuid.uuid4()}{os.path.splitext(archivo.name)[1]}"
            categoría.url_imagen = guardar_archivo_hibrido(nombre_unico, archivo)
        else:
            categoría.url_imagen = petición.POST.get('url_imagen', categoría.url_imagen).strip()
        categoría.save()
        messages.success(petición, f'Categoría "{categoría.nombre}" actualizada.')
        return redirect('/admin-panel/categorias/')
    return render(petición, 'admin_panel/category_form.html', {
        'categoria': categoría,
        'titulo_pagina': f'Cat: {categoría.nombre} — MIKITECH',
    })


@requerir_administrador
@require_POST
def eliminar_categoria(petición, id_cat):
    """Remover categorías si están vacías o bajo confirmación."""
    categoría = get_object_or_404(Categoria, id=id_cat)
    try:
        nombre = categoría.nombre
        categoría.delete()
        messages.success(petición, f'Categoría "{nombre}" removida.')
    except Exception as e:
        messages.error(petición, f'Fallo en borrado: {str(e)}.')
    
    return redirect('/admin-panel/categorias/')


@requerir_administrador
def gestion_usuarios(petición):
    """Monitorización de perfiles de clientes y administradores."""
    usuarios = Perfil.objects.prefetch_related(
        'pedidos__detalles__producto',
        'reseñas__producto'
    ).all().order_by('-creado_el')
    conteo_admin = usuarios.filter(rol='admin').count()
    conteo_cliente = usuarios.filter(rol='client').count()
    return render(petición, 'admin_panel/users.html', {
        'usuarios': usuarios,
        'conteo_admin': conteo_admin,
        'conteo_cliente': conteo_cliente,
        'titulo_pagina': 'Gestión de Usuarios — MIKITECH',
    })


@requerir_administrador
def editar_usuario(petición, id_usuario):
    """Permite al administrador editar cualquier usuario (datos de texto y foto)."""
    usuario = get_object_or_404(Perfil, id=id_usuario)
    
    if petición.method == 'POST':
        from users.photo_manager import UserPhotoManager
        
        nombre_completo = petición.POST.get('nombre_completo', '').strip()
        nombre_usuario = petición.POST.get('nombre_usuario', '').strip()
        email = petición.POST.get('email', '').strip()
        rol = petición.POST.get('rol', '').strip()
        esta_activo = petición.POST.get('esta_activo') == 'on'
        telefono = petición.POST.get('telefono', '').strip()
        biografia = petición.POST.get('biografia', '').strip()
        ciudad = petición.POST.get('ciudad', '').strip()
        pais = petición.POST.get('pais', '').strip()
        
        # Validar campos requeridos
        if not nombre_usuario:
            messages.error(petición, "El nombre de usuario es obligatorio.")
            return render(petición, 'admin_panel/edit_user.html', {
                'usuario': usuario,
                'titulo_pagina': f'Editar Usuario: {usuario.nombre_mostrado} — MIKITECH',
            })
            
        # Comprobar si el nombre de usuario ya existe en otro perfil
        if Perfil.objects.filter(nombre_usuario=nombre_usuario).exclude(id=usuario.id).exists():
            messages.error(petición, "El nombre de usuario ya está en uso.")
            return render(petición, 'admin_panel/edit_user.html', {
                'usuario': usuario,
                'titulo_pagina': f'Editar Usuario: {usuario.nombre_mostrado} — MIKITECH',
            })
            
        data = {
            'nombre_completo': nombre_completo,
            'nombre_usuario': nombre_usuario,
            'rol': rol,
            'esta_activo': esta_activo,
            'telefono': telefono,
            'biografia': biografia,
            'ciudad': ciudad,
            'pais': pais,
        }
        
        if email:
            data['email'] = email
            
        photo_file = petición.FILES.get('avatar')
        
        updated_profile, error = UserPhotoManager.update_user(id_usuario, data, photo_file)
        
        if error:
            messages.error(petición, f"Error actualizando el usuario: {error}")
        else:
            messages.success(petición, f"Usuario '{updated_profile.nombre_mostrado}' actualizado correctamente.")
            return redirect('/admin-panel/usuarios/')
            
    return render(petición, 'admin_panel/edit_user.html', {
        'usuario': usuario,
        'titulo_pagina': f'Editar Usuario: {usuario.nombre_mostrado} — MIKITECH',
    })


@requerir_administrador
def moderacion_resenas(petición):
    """Panel de moderación para comentarios de productos."""
    reseñas = Reseña.objects.select_related('producto', 'usuario').order_by('-creado_el')
    return render(petición, 'admin_panel/reviews.html', {
        'resenas': reseñas,
        'titulo_pagina': 'Moderación de Reseñas — MIKITECH',
    })


@requerir_administrador
@require_POST
def eliminar_resena(petición, id_resena):
    """Eliminar un comentario inapropiado o falso."""
    reseña = get_object_or_404(Reseña, id=id_resena)
    reseña.delete()
    messages.success(petición, 'Reseña eliminada con éxito.')
    return redirect('/admin-panel/resenas/')


@requerir_administrador
@require_POST
def enviar_notificacion_resena(petición, id_resena):
    """Enviar un aviso de moderación al usuario de la reseña."""
    reseña = get_object_or_404(Reseña, id=id_resena)
    mensaje = petición.POST.get('mensaje', '').strip()
    
    if not mensaje:
        messages.error(petición, 'Debes escribir un mensaje para la notificación.')
        return redirect('/admin-panel/resenas/')
    
    try:
        Notificacion.objects.create(
            id=uuid.uuid4(),
            usuario=reseña.usuario,
            mensaje=f"Aviso de Moderación MIKITECH: {mensaje} (Ref: {reseña.producto.nombre})",
        )
        messages.success(petición, f'Notificación enviada a {reseña.usuario.nombre_usuario}.')
    except Exception as e:
        messages.error(petición, f'Error al enviar notificación: {str(e)}')
        
    return redirect('/admin-panel/resenas/')


@requerir_administrador
def reportes_dashboard(petición):
    """Métricas globales y reportes exportables con filtros de tiempo dinámicos."""
    import datetime
    from django.utils import timezone
    from django.db.models import Sum, Count, Avg, F
    from django.db.models.functions import TruncMonth
    
    dias = petición.GET.get('dias', 'all')
    filtro_fecha = {}
    filtro_fecha_usuario = {}
    
    if dias.isdigit():
        fecha_limite = timezone.now() - datetime.timedelta(days=int(dias))
        filtro_fecha = {'creado_el__gte': fecha_limite}
        filtro_fecha_usuario = {'creado_el__gte': fecha_limite}

    # Estadísticas base
    estadísticas = {
        'inventario_total': Producto.objects.count(), 
        'nuevos_productos': Producto.objects.filter(**filtro_fecha).count(), 
        'total_usuarios': Perfil.objects.count(),
        'total_resenas': Reseña.objects.filter(**filtro_fecha).count(),
        'pedidos_total': Pedido.objects.filter(**filtro_fecha).count(),
        'filtro_actual': dias,
    }
    
    # 1. Ticket Promedio (AOV) e Ingresos Totales
    pedidos_exitosos = Pedido.objects.filter(estado__in=['delivered', 'shipped', 'Entregado', 'Enviado'])
    pedidos_periodo = pedidos_exitosos.filter(**filtro_fecha)
    
    avg_order = pedidos_periodo.aggregate(promedio=Avg('monto_total'))['promedio'] or 0
    total_ingresos = pedidos_periodo.aggregate(total=Sum('monto_total'))['total'] or 0
    
    estadísticas['ticket_promedio'] = avg_order
    estadísticas['ingresos_periodo'] = total_ingresos

    # 2. Datos para el Gráfico (Últimos 6 meses sin GAPS)
    chart_labels = []
    chart_data = []
    
    meses_nombres = {
        1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
        7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
    }

    hoy = timezone.now()
    datos_por_mes = {}
    
    # Obtener datos de la DB
    seis_meses_atras = hoy - datetime.timedelta(days=180)
    ventas_db = pedidos_exitosos.filter(creado_el__gte=seis_meses_atras) \
        .annotate(mes_idx=TruncMonth('creado_el')) \
        .values('mes_idx') \
        .annotate(total=Sum('monto_total'))
    
    for v in ventas_db:
        periodo_key = v['mes_idx'].strftime("%Y-%m")
        datos_por_mes[periodo_key] = float(v['total']) / 1_000_000

    # Llenar la serie (6 meses atrás hasta hoy)
    for i in range(5, -1, -1):
        # Usamos aproximación de 30 días para retroceder meses de forma segura
        fecha_eval = hoy - datetime.timedelta(days=i*30)
        # Ajustamos al primer día del mes para la clave de búsqueda
        primer_dia_mes = fecha_eval.replace(day=1)
        key = primer_dia_mes.strftime("%Y-%m")
        chart_labels.append(meses_nombres[primer_dia_mes.month])
        chart_data.append(datos_por_mes.get(key, 0))

    pedidos = Pedido.objects.filter(**filtro_fecha).select_related('usuario').prefetch_related('detalles__producto').order_by('-creado_el')
    
    return render(petición, 'admin_panel/reports.html', {
        'estadisticas': estadísticas,
        'pedidos': pedidos,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'titulo_pagina': 'Reportes y Estadísticas — MIKITECH',
    })


@requerir_administrador
def gestion_logistica(petición):
    """Módulo de despacho y entrega de pedidos."""
    estado_filtro_es = petición.GET.get('estado', 'all')
    
    # Mapeo invertido para la consulta en DB
    mapping_inverso = {
        'Pendiente': 'pending',
        'Enviado': 'shipped',
        'Entregado': 'delivered',
        'Cancelado': 'cancelled'
    }
    
    db_status = mapping_inverso.get(estado_filtro_es, estado_filtro_es)

    if db_status != 'all':
        pedidos = Pedido.objects.filter(estado=db_status).select_related('usuario', 'repartidor').order_by('-creado_el')
    else:
        pedidos = Pedido.objects.all().select_related('usuario', 'repartidor').order_by('-creado_el')

    # Estadísticas rápidas para la vista (Usando términos en inglés de la DB)
    stats = {
        'pendientes': Pedido.objects.filter(estado__in=['pending', 'processing']).count(),
        'en_camino': Pedido.objects.filter(estado='shipped').count(),
        'entregados': Pedido.objects.filter(estado='delivered').count(),
        'sin_repartidor': Pedido.objects.filter(repartidor__isnull=True, estado__in=['pending', 'processing', 'shipped']).count(),
        'filtro_actual': estado_filtro_es,
    }

    # Cargar repartidores para el modal de asignación
    repartidores = Perfil.objects.filter(rol='repartidor', esta_activo=True).order_by('nombre_usuario')

    return render(petición, 'admin_panel/logistics.html', {
        'pedidos': pedidos,
        'stats': stats,
        'repartidores': repartidores,
        'titulo_pagina': 'Logística y Despacho — MIKITECH',
    })


@requerir_administrador
@require_POST
def asignar_repartidor_admin(petición, id_pedido):
    """Asignación manual de un piloto a un pedido."""
    pedido = get_object_or_404(Pedido, id=id_pedido)
    id_repartidor = petición.POST.get('id_repartidor')
    
    if id_repartidor:
        repartidor = get_object_or_404(Perfil, id=id_repartidor, rol='repartidor')
        pedido.repartidor = repartidor
        pedido.save()
        
        # Notificar al cliente sobre la asignación del repartidor
        from users.models import Notificacion
        Notificacion.objects.create(
            id=uuid.uuid4(),
            usuario=pedido.usuario,
            mensaje=f"[Repartidor Asignado] ¡Hola! Tu pedido #{str(pedido.id)[:8]} ya tiene un piloto asignado: {repartidor.nombre_usuario}. Estamos preparando tu despacho.",
        )
        
        messages.success(petición, f'Pedido #{str(pedido.id)[:8]} asignado a {repartidor.nombre_usuario}. Cliente notificado.')
    else:
        messages.warning(petición, 'No se seleccionó ningún repartidor.')
        
    referer = petición.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    return redirect('admin_logistics')


@requerir_administrador
@require_POST
def cambiar_estado_pedido(petición, id_pedido):
    """Actualiza el estado de un pedido y notifica al cliente si es necesario."""
    pedido = get_object_or_404(Pedido, id=id_pedido)
    if petición.method == 'POST':
        nuevo_estado_es = petición.POST.get('nuevo_estado')
        
        # Mapeo de español (UI) a inglés (DB) para cumplir con el CHECK constraint
        mapping = {
            'Pendiente': 'pending',
            'Enviado': 'shipped',
            'Entregado': 'delivered',
            'Cancelado': 'cancelled'
        }
        
        status_to_save = mapping.get(nuevo_estado_es)
        
        if status_to_save:
            if status_to_save == 'delivered':
                cedula_cliente = petición.POST.get('cedula_cliente', '').strip()
                if not cedula_cliente or cedula_cliente != pedido.cedula:
                    messages.error(petición, f"Validación de Cédula Fallida: La cédula ingresada no coincide.")
                    referer = petición.META.get('HTTP_REFERER')
                    if referer:
                        return redirect(referer)
                    return redirect('admin_orders')
                    
            estado_anterior = pedido.estado
            pedido.estado = status_to_save
            if status_to_save == 'delivered':
                from django.utils import timezone
                pedido.entregado_el = timezone.now()
            pedido.save()
            
            # Notificaciones Automáticas según el estado
            from users.models import Notificacion
            notif_msg = None
            
            if status_to_save == 'shipped' and estado_anterior != 'shipped':
                notif_msg = f"[En Camino] 🚚 ¡Grandes noticias! Tu pedido #{str(pedido.id)[:8]} ya ha sido despachado y está en manos de nuestro repartidor. ¡Llegará pronto!"
            elif status_to_save == 'delivered' and estado_anterior != 'delivered':
                notif_msg = f"[Entregado] ✅ Tu pedido #{str(pedido.id)[:8]} ha sido entregado exitosamente. ¡Gracias por confiar en MIKITECH!"
            elif status_to_save == 'cancelled' and estado_anterior != 'cancelled':
                notif_msg = f"[Cancelado] ❌ Tu pedido #{str(pedido.id)[:8]} ha sido cancelado. Si tienes dudas, contáctanos a soporte."
                
            if notif_msg:
                Notificacion.objects.create(
                    id=uuid.uuid4(),
                    usuario=pedido.usuario,
                    mensaje=notif_msg,
                )
                messages.success(petición, f"Pedido #{str(pedido.id)[:8]} actualizado a {nuevo_estado_es.upper()}. Cliente notificado.")
            else:
                messages.success(petición, f"Estado del pedido #{str(pedido.id)[:8]} actualizado a {nuevo_estado_es}.")
            
    referer = petición.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    return redirect('admin_logistics')


@requerir_administrador
@xframe_options_sameorigin
def ver_factura_pedido(petición, id_pedido):
    """Vista detallada de factura para impresión administrativa o datos JSON."""
    pedido = get_object_or_404(Pedido.objects.prefetch_related('detalles__producto'), id=id_pedido)
    perfil = pedido.usuario

    if petición.GET.get('format') == 'json':
        from django.http import JsonResponse
        detalles_list = []
        for det in pedido.detalles.all():
            detalles_list.append({
                'nombre': det.producto.nombre if det.producto else "Producto descatalogado",
                'cantidad': det.cantidad,
                'precio': float(det.precio_unitario)
            })
        
        # Localized/formatted date
        fecha_str = pedido.creado_el.strftime('%d/%m/%Y') if pedido.creado_el else ''
        
        return JsonResponse({
            'order_id': str(pedido.id),
            'creado_el': fecha_str,
            'monto_total': float(pedido.monto_total),
            'cliente_nombre': perfil.nombre_mostrado if perfil else "Cliente Genérico",
            'cliente_username': perfil.nombre_usuario if perfil else "cliente",
            'cliente_telefono': pedido.telefono or (perfil.telefono if perfil and perfil.telefono else '') or '+57 300 000 0000',
            'direccion_envio': pedido.direccion_envio or 'Retiro en tienda',
            'cedula': pedido.cedula or 'N/A',
            'detalles': detalles_list
        })
    
    # Cálculos de impuestos para la factura legal
    import decimal
    iva_porcentaje = decimal.Decimal('0.19')
    subtotal = pedido.monto_total / (decimal.Decimal('1') + iva_porcentaje)
    iva_monto = pedido.monto_total - subtotal

    return render(petición, 'admin_panel/invoice.html', {
        'pedido': pedido,
        'detalles': pedido.detalles.all(),
        'perfil': perfil,
        'subtotal': subtotal,
        'iva_monto': iva_monto,
        'titulo_pagina': f'Factura #{str(pedido.id)[:8].upper()} — MIKITECH',
    })


@requerir_administrador
def descargar_plantilla_excel(petición):
    import openpyxl
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Productos"
    columnas = ["NOMBRE", "CATEGORIA", "PRECIO", "STOCK", "MARCA", "DESCRIPCION", "URL_IMAGEN"]
    ws.append(columnas)
    ws.append(["Audífonos Bluetooth Pro", "Electrónica", 150000, 50, "Sony", "Audífonos con cancelación de ruido", "https://ejemplo.com/imagen.jpg"])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_productos_mikitech.xlsx"'
    wb.save(response)
    return response


@requerir_administrador
def carga_masiva_productos(petición):
    import openpyxl
    import json as json_lib
    campos_referencia = [
        {'nombre': 'NOMBRE', 'requerido': True, 'descripcion': 'Nombre del producto.'},
        {'nombre': 'CATEGORIA', 'requerido': True, 'descripcion': 'Nombre de la categoría (se crea si no existe).'},
        {'nombre': 'PRECIO', 'requerido': True, 'descripcion': 'Precio de venta al público sin puntos ni comas.'},
        {'nombre': 'STOCK', 'requerido': True, 'descripcion': 'Cantidad disponible en bodega.'},
        {'nombre': 'MARCA', 'requerido': False, 'descripcion': 'Marca del fabricante.'},
        {'nombre': 'DESCRIPCION', 'requerido': False, 'descripcion': 'Detalle completo del producto.'},
        {'nombre': 'URL_IMAGEN', 'requerido': False, 'descripcion': 'Enlace directo a la imagen principal.'},
    ]

    reporte = None
    
    if petición.method == 'POST':
        archivo = petición.FILES.get('archivo_excel')
        omitir_errores = petición.POST.get('omitir_errores') == 'on'
        omitir_duplicados = petición.POST.get('omitir_duplicados') == 'on'
        
        if not archivo:
            messages.error(petición, 'Por favor sube un archivo válido (.xlsx o .json).')
        elif not (archivo.name.endswith('.xlsx') or archivo.name.endswith('.json')):
            messages.error(petición, 'Formato no soportado. Usa archivos .xlsx o .json.')
        else:
            try:
                from django.utils.text import slugify
                
                creados = 0
                errores = 0
                omitidos = 0
                detalles = []
                
                # ── Construir lista unificada de filas ──────────────────────
                filas = []  # lista de dicts con claves normalizadas
                
                if archivo.name.endswith('.json'):
                    # ── Lectura JSON ────────────────────────────────────────
                    contenido = archivo.read().decode('utf-8')
                    datos_json = json_lib.loads(contenido)
                    if not isinstance(datos_json, list):
                        raise Exception("El JSON debe ser un arreglo de objetos. Ej: [{...}, {...}]")
                    for i, obj in enumerate(datos_json):
                        if not isinstance(obj, dict):
                            raise Exception(f"Elemento #{i+1} no es un objeto válido.")
                        # Normalizar claves a MAYÚSCULAS
                        norm = {k.strip().upper(): v for k, v in obj.items()}
                        filas.append((i + 1, norm))
                else:
                    # ── Lectura Excel ───────────────────────────────────────
                    wb = openpyxl.load_workbook(archivo)
                    ws = wb.active
                    headers = [str(cell.value).strip().upper() if cell.value else "" for cell in ws[1]]
                    col_map = {name: i for i, name in enumerate(headers)}
                    
                    for fila_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        if not any(row):
                            continue
                        norm = {}
                        for col_name, col_i in col_map.items():
                            if col_i < len(row):
                                norm[col_name] = row[col_i]
                        filas.append((fila_idx, norm))
                
                # ── Verificar columnas requeridas ──────────────────────────
                if filas:
                    req_cols = ["NOMBRE", "CATEGORIA", "PRECIO", "STOCK"]
                    primera_fila_keys = set(filas[0][1].keys())
                    faltantes = [c for c in req_cols if c not in primera_fila_keys]
                    if faltantes:
                        raise Exception(f"El archivo debe contener los campos: {', '.join(faltantes)}")
                
                # ── Procesar cada fila ─────────────────────────────────────
                for fila_idx, datos in filas:
                    nombre = datos.get("NOMBRE")
                    cat_nombre = datos.get("CATEGORIA")
                    precio = datos.get("PRECIO")
                    stock = datos.get("STOCK")
                    marca = datos.get("MARCA", "")
                    desc = datos.get("DESCRIPCION", "")
                    url_img = datos.get("URL_IMAGEN", "")
                    
                    if not nombre or not cat_nombre or precio is None or stock is None:
                        errores += 1
                        detalles.append({"fila": fila_idx, "nombre": nombre, "categoria": cat_nombre, "estado": "error", "mensaje": "Faltan datos obligatorios"})
                        if not omitir_errores: break
                        continue
                        
                    if omitir_duplicados and Producto.objects.filter(nombre=nombre).exists():
                        omitidos += 1
                        detalles.append({"fila": fila_idx, "nombre": nombre, "categoria": cat_nombre, "estado": "omitido", "mensaje": "Producto duplicado"})
                        continue
                        
                    try:
                        cat_slug = slugify(cat_nombre)
                        categoria, _ = Categoria.objects.get_or_create(
                            enlace=cat_slug, 
                            defaults={"nombre": cat_nombre, "id": uuid.uuid4()}
                        )
                        
                        prod_slug = slugify(nombre)
                        base_enlace = prod_slug
                        contador = 1
                        while Producto.objects.filter(enlace=prod_slug).exists():
                            prod_slug = f"{base_enlace}-{contador}"
                            contador += 1
                            
                        Producto.objects.create(
                            id=uuid.uuid4(),
                            categoria=categoria,
                            nombre=nombre,
                            enlace=prod_slug,
                            precio=float(precio),
                            existencias=int(stock),
                            marca=marca or "",
                            descripcion=desc or "",
                            url_imagen_principal=url_img or "",
                            esta_activo=True
                        )
                        creados += 1
                        detalles.append({"fila": fila_idx, "nombre": nombre, "categoria": cat_nombre, "estado": "ok", "mensaje": "Producto creado con éxito"})
                    except Exception as e:
                        errores += 1
                        detalles.append({"fila": fila_idx, "nombre": nombre, "categoria": cat_nombre, "estado": "error", "mensaje": str(e)[:50]})
                        if not omitir_errores: break
                
                reporte = {
                    "creados": creados,
                    "errores": errores,
                    "omitidos": omitidos,
                    "detalles": detalles
                }
                messages.success(petición, 'Proceso de importación finalizado.')
            except json_lib.JSONDecodeError as e:
                messages.error(petición, f'Error de sintaxis en el JSON: {str(e)[:80]}')
            except Exception as e:
                messages.error(petición, f'Error al procesar el archivo: {str(e)}')
            
    return render(petición, 'admin_panel/bulk_upload.html', {
        'campos_referencia': campos_referencia,
        'reporte': reporte,
        'titulo_pagina': 'Carga Masiva — MIKITECH'
    })


@requerir_administrador
def gestion_pedidos(petición):
    """Gestión interactiva de pedidos (pedidos en proceso e historial)."""
    pedidos = Pedido.objects.select_related('usuario', 'repartidor').prefetch_related('detalles__producto').order_by('-creado_el')
    repartidores = Perfil.objects.filter(rol='repartidor', esta_activo=True).order_by('nombre_usuario')
    
    return render(petición, 'admin_panel/orders.html', {
        'pedidos': pedidos,
        'repartidores': repartidores,
        'titulo_pagina': 'Gestión de Pedidos — MIKITECH',
    })


def asegurar_notificaciones_admin(usuario_id):
    """
    Genera dinámicamente notificaciones para el administrador si existen:
    - Productos con bajo stock (<= 5).
    - Reseñas pendientes de aprobación.
    - Pedidos pendientes.
    """
    from products.models import Producto
    from interactions.models import Reseña, Pedido
    from users.models import Notificacion
    
    # 1. Bajo stock
    productos_bajo_stock = Producto.objects.filter(existencias__lte=5, esta_activo=True)
    for p in productos_bajo_stock:
        mensaje = f"⚠️ Bajo stock en {p.nombre} ({p.existencias} unidades). Reabastece aquí.|/admin-panel/productos/editar/{p.id}/"
        if not Notificacion.objects.filter(usuario_id=usuario_id, mensaje__startswith=f"⚠️ Bajo stock en {p.nombre}", esta_leida=False).exists():
            Notificacion.objects.create(
                usuario_id=usuario_id,
                mensaje=mensaje,
                esta_leida=False
            )
            
    # 2. Reseñas pendientes
    resenas_pendientes = Reseña.objects.filter(esta_aprobada=False)
    for r in resenas_pendientes:
        mensaje = f"💬 Nueva reseña pendiente de moderación para {r.producto.nombre}.|/admin-panel/resenas/"
        if not Notificacion.objects.filter(usuario_id=usuario_id, mensaje__startswith="💬 Nueva reseña pendiente de moderación", esta_leida=False).exists():
            Notificacion.objects.create(
                usuario_id=usuario_id,
                mensaje=mensaje,
                esta_leida=False
            )
            
    # 3. Pedidos pendientes
    pedidos_pendientes = Pedido.objects.filter(estado__in=['pending', 'processing', 'Pendiente'])
    for ped in pedidos_pendientes:
        id_corto = str(ped.id)[:8]
        mensaje = f"📦 Pedido pendiente #{id_corto} esperando asignación de logística.|/admin-panel/logistica/"
        if not Notificacion.objects.filter(usuario_id=usuario_id, mensaje__startswith=f"📦 Pedido pendiente #{id_corto}", esta_leida=False).exists():
            Notificacion.objects.create(
                usuario_id=usuario_id,
                mensaje=mensaje,
                esta_leida=False
            )


@requerir_administrador
def leer_notificacion_admin(petición, id_notificacion):
    """Marca una notificación específica como leída y redirige al destino."""
    notificacion = get_object_or_404(Notificacion, id=id_notificacion)
    notificacion.esta_leida = True
    notificacion.save()
    
    mensaje = notificacion.mensaje
    if '|' in mensaje:
        partes = mensaje.split('|')
        url_destino = partes[1].strip()
        return redirect(url_destino)
        
    return redirect('/admin-panel/')


# -------------------------------------------------------------
# SISTEMA DE INVITACIONES PARA ADMINISTRADORES & SEGURIDAD RBAC
# -------------------------------------------------------------

@requerir_administrador
@requiere_permiso('gestionar_usuarios')
def gestion_invitaciones_admin(petición):
    """Muestra todas las invitaciones y los administradores activos."""
    invitaciones = InvitacionAdmin.objects.all().order_by('-fecha_envio')
    
    # Actualizar estado de expiración de invitaciones pendientes si ya pasó su fecha
    from django.utils import timezone
    ahora = timezone.now()
    for inv in invitaciones.filter(estado='pendiente', fecha_expiracion__lt=ahora):
        inv.estado = 'expirada'
        inv.save()
        
    usuarios = Perfil.objects.filter(rol='admin').order_by('nombre_usuario')
    
    return render(petición, 'admin_panel/invitations.html', {
        'titulo_pagina': 'Gestionar Administradores — MIKITECH',
        'invitaciones': invitaciones,
        'usuarios_admin': usuarios,
    })


@requerir_administrador
@requiere_permiso('gestionar_usuarios')
def crear_invitacion_admin(petición):
    """Crea una nueva invitación y registra al administrador temporalmente."""
    if petición.method == 'POST':
        email = petición.POST.get('email', '').strip()
        nombre_completo = petición.POST.get('nombre_completo', '').strip()
        notas = petición.POST.get('notas_internas', '').strip()

        if not email:
            messages.error(petición, 'El correo electrónico es obligatorio.')
            return redirect('/admin-panel/invitaciones/')

        import re
        if not re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email):
            messages.error(petición, 'Formato de correo electrónico inválido.')
            return redirect('/admin-panel/invitaciones/')

        # Verificar si el email ya existe en perfiles o invitaciones activas
        from django.db import connection
        table_name = '"auth.users"' if connection.vendor == 'sqlite' else 'auth.users'
        try:
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT id FROM {table_name} WHERE email = %s", [email])
                if cursor.fetchone():
                    messages.error(petición, 'Este correo electrónico ya está registrado en el sistema.')
                    return redirect('/admin-panel/invitaciones/')
        except Exception as e:
            print(f"[crear_invitacion_admin] Error verificando duplicidad: {e}")

        if InvitacionAdmin.objects.filter(email=email, estado='pendiente').exists():
            messages.error(petición, 'Ya existe una invitación pendiente para este correo electrónico.')
            return redirect('/admin-panel/invitaciones/')

        # Generar nombre de usuario único y limpio
        import random
        cleaned_name = re.sub(r'[^a-zA-Z]', '', nombre_completo.lower().split()[0]) if nombre_completo else 'admin'
        cleaned_last = re.sub(r'[^a-zA-Z]', '', nombre_completo.lower().split()[1][0]) if nombre_completo and len(nombre_completo.split()) > 1 else ''
        username_prefix = f"admin_{cleaned_name}{cleaned_last}"
        
        while True:
            rand_num = random.randint(1000, 9999)
            usuario_generado = f"{username_prefix}_{rand_num}"
            if not Perfil.objects.filter(nombre_usuario=usuario_generado).exists() and not InvitacionAdmin.objects.filter(usuario_generado=usuario_generado).exists():
                break

        # Generar contraseña temporal segura
        import secrets
        import string
        alphabet = string.ascii_letters + string.digits + "@$!%*?&"
        while True:
            password_temporal = ''.join(secrets.choice(alphabet) for _ in range(12))
            if (any(c.isupper() for c in password_temporal) and
                any(c.islower() for c in password_temporal) and
                any(c.isdigit() for c in password_temporal) and
                any(c in "@$!%*?&" for c in password_temporal)):
                break

        # Registrar en Supabase Auth / SQLite
        from users.supabase_auth import registrar_usuario, registrar_usuario_sql
        datos, error = registrar_usuario(email, password_temporal, nombre_completo, usuario_generado, rol='admin')
        if error and "confirmation email" in error.lower():
            datos, error = registrar_usuario_sql(email, password_temporal, nombre_completo, usuario_generado, rol='admin')

        if error:
            messages.error(petición, f'Error al registrar el usuario en autenticación: {error}')
            return redirect('/admin-panel/invitaciones/')

        id_usuario = datos.get('user', {}).get('id')

        # Crear invitación
        from django.utils import timezone
        from django.contrib.auth.hashers import make_password
        expires_at = timezone.now() + timezone.timedelta(days=7)
        
        invitacion = InvitacionAdmin.objects.create(
            email=email,
            nombre_completo=nombre_completo,
            usuario_generado=usuario_generado,
            password_temporal_hash=make_password(password_temporal),
            fecha_expiracion=expires_at,
            estado='pendiente',
            notas_internas=notas,
            creado_por=petición.perfil_usuario
        )

        # Configurar perfil
        perfil, _ = Perfil.objects.update_or_create(
            id=id_usuario,
            defaults={
                'nombre_completo': nombre_completo,
                'nombre_usuario': usuario_generado,
                'rol_rbac': Rol.objects.get(codigo='admin'),
                'rol': 'admin',
                'invitacion': invitacion,
                'password_cambiada': False,
                'esta_activo': True
            }
        )

        # Enviar correo de invitación
        from django.core.mail import send_mail
        subject = "[MIKITECH] Has sido invitado como Administrador"
        url_login = petición.build_absolute_uri('/admin-panel/login/')
        message = f"""Hola {nombre_completo or 'Administrador'},

Has sido invitado para ser administrador en el Panel de Control de MIKITECH.

Tus credenciales de acceso temporal son:
- Usuario / Correo: {email}
- Nombre de usuario: {usuario_generado}
- Contraseña temporal: {password_temporal}

⚠️ Por motivos de seguridad, esta contraseña es temporal y caducará en 7 días ({expires_at.strftime('%Y-%m-%d')}). Deberás cambiarla obligatoriamente en tu primer inicio de sesión.

Puedes iniciar sesión aquí: {url_login}

Si tienes problemas, contacta al superadministrador de MIKITECH.
"""
        try:
            send_mail(subject, message, settings.DEFAULT_FROM_EMAIL or 'noreply@mikitech.com', [email])
            messages.success(petición, f'✅ Invitación enviada exitosamente a {email} con usuario {usuario_generado}.')
        except Exception as e:
            messages.warning(petición, f'⚠️ Invitación creada pero el correo no pudo enviarse: {e}. Credenciales: {usuario_generado} / {password_temporal}')

    return redirect('/admin-panel/invitaciones/')


@requerir_administrador
@requiere_permiso('gestionar_usuarios')
@require_POST
def revocar_invitacion_admin(petición, id_invitacion):
    """Revoca una invitación pendiente y desactiva el perfil del usuario."""
    invitacion = get_object_or_404(InvitacionAdmin, id=id_invitacion)
    if invitacion.estado == 'pendiente':
        invitacion.estado = 'revocada'
        invitacion.save()
        
        # Desactivar perfil de usuario asociado
        Perfil.objects.filter(invitacion=invitacion).update(esta_activo=False)
        messages.success(petición, f'Invitación para {invitacion.email} revocada exitosamente.')
    else:
        messages.error(petición, 'Solo se pueden revocar invitaciones pendientes.')
        
    return redirect('/admin-panel/invitaciones/')


@requerir_administrador
@requiere_permiso('gestionar_usuarios')
@require_POST
def reenviar_invitacion_admin(petición, id_invitacion):
    """Reenvía la invitación restableciendo la contraseña del usuario y fecha de expiración."""
    invitacion = get_object_or_404(InvitacionAdmin, id=id_invitacion)
    
    # 1. Generar contraseña temporal segura
    import secrets
    import string
    alphabet = string.ascii_letters + string.digits + "@$!%*?&"
    while True:
        password_temporal = ''.join(secrets.choice(alphabet) for _ in range(12))
        if (any(c.isupper() for c in password_temporal) and
            any(c.islower() for c in password_temporal) and
            any(c.isdigit() for c in password_temporal) and
            any(c in "@$!%*?&" for c in password_temporal)):
            break

    # 2. Buscar perfil existente para eliminarlo y recrearlo en autenticación (método seguro y libre de permisos)
    perfil_invitado = Perfil.objects.filter(invitacion=invitacion).first()
    if not perfil_invitado:
        messages.error(petición, 'No se encontró un perfil asociado a esta invitación.')
        return redirect('/admin-panel/invitaciones/')

    from django.db import connection
    table_name = '"auth.users"' if connection.vendor == 'sqlite' else 'auth.users'
    user_id = str(perfil_invitado.id)
    
    try:
        # Borrar registros para recrearlos
        perfil_invitado.delete()
        with connection.cursor() as cursor:
            cursor.execute(f"DELETE FROM {table_name} WHERE id = %s", [user_id])
    except Exception as e:
        print(f"[reenviar_invitacion_admin] Error al limpiar registros previos: {e}")

    # 3. Registrar de nuevo en Auth
    from users.supabase_auth import registrar_usuario, registrar_usuario_sql
    datos, error = registrar_usuario(invitacion.email, password_temporal, invitacion.nombre_completo, invitacion.usuario_generado, rol='admin')
    if error and "confirmation email" in error.lower():
        datos, error = registrar_usuario_sql(invitacion.email, password_temporal, invitacion.nombre_completo, invitacion.usuario_generado, rol='admin')

    if error:
        messages.error(petición, f'Error al registrar el usuario en autenticación al reenviar: {error}')
        return redirect('/admin-panel/invitaciones/')

    nuevo_id_usuario = datos.get('user', {}).get('id')

    # 4. Actualizar invitación
    from django.utils import timezone
    from django.contrib.auth.hashers import make_password
    expires_at = timezone.now() + timezone.timedelta(days=7)
    
    invitacion.password_temporal_hash = make_password(password_temporal)
    invitacion.fecha_expiracion = expires_at
    invitacion.estado = 'pendiente'
    invitacion.save()

    # 5. Volver a configurar perfil
    perfil, _ = Perfil.objects.update_or_create(
        id=nuevo_id_usuario,
        defaults={
            'nombre_completo': invitacion.nombre_completo,
            'nombre_usuario': invitacion.usuario_generado,
            'rol_rbac': Rol.objects.get(codigo='admin'),
            'rol': 'admin',
            'invitacion': invitacion,
            'password_cambiada': False,
            'esta_activo': True
        }
    )

    # 6. Enviar correo de invitación
    from django.core.mail import send_mail
    subject = "[MIKITECH] Reenvío de Invitación de Administrador"
    url_login = petición.build_absolute_uri('/admin-panel/login/')
    message = f"""Hola {invitacion.nombre_completo or 'Administrador'},

Se ha reenviado tu invitación para ser administrador en el Panel de Control de MIKITECH.

Tus nuevas credenciales de acceso temporal son:
- Usuario / Correo: {invitacion.email}
- Nombre de usuario: {invitacion.usuario_generado}
- Contraseña temporal: {password_temporal}

⚠️ Esta contraseña es temporal y caducará en 7 días ({expires_at.strftime('%Y-%m-%d')}). Deberás cambiarla obligatoriamente en tu primer inicio de sesión.

Puedes iniciar sesión aquí: {url_login}
"""
    try:
        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL or 'noreply@mikitech.com', [invitacion.email])
        messages.success(petición, f'✅ Invitación reenviada exitosamente a {invitacion.email}.')
    except Exception as e:
        messages.warning(petición, f'⚠️ Invitación actualizada pero el correo no pudo enviarse: {e}. Credenciales: {invitacion.usuario_generado} / {password_temporal}')

    return redirect('/admin-panel/invitaciones/')


@requerir_administrador
@requiere_permiso('gestionar_usuarios')
@require_POST
def desactivar_usuario_admin(petición, id_usuario):
    """Desactiva a un administrador activo para impedir su inicio de sesión."""
    usuario = get_object_or_404(Perfil, id=id_usuario)
    if usuario.rol == 'admin':
        if str(usuario.id) == str(petición.session.get('usuario_id')):
            messages.error(petición, 'No puedes desactivar tu propio usuario.')
        else:
            usuario.esta_activo = False
            usuario.save()
            messages.success(petición, f'Administrador {usuario.nombre_usuario} desactivado exitosamente.')
    else:
        messages.error(petición, 'El usuario seleccionado no es un administrador.')
        
    return redirect('/admin-panel/invitaciones/')


def cambiar_contrasena_forzado(petición):
    """Obliga al administrador invitado a cambiar su contraseña temporal."""
    # Verificar que el usuario esté logueado
    usuario_id = petición.session.get('usuario_id')
    if not usuario_id:
        return redirect('/admin-panel/login/')
        
    # Obtener el perfil
    perfil = get_object_or_404(Perfil, id=usuario_id)
    
    # Solo permitir acceso si no ha cambiado la contraseña temporal
    if perfil.password_cambiada:
        return redirect('/admin-panel/')
        
    contexto = {
        'titulo_pagina': 'Cambiar Contraseña Obligatorio — MIKITECH',
        'perfil': perfil,
    }
    
    if petición.method == 'POST':
        nueva_clave = petición.POST.get('nueva_clave', '')
        confirmar_clave = petición.POST.get('confirmar_clave', '')
        
        if not nueva_clave or not confirmar_clave:
            contexto['error'] = 'Por favor completa todos los campos del formulario.'
            return render(petición, 'admin_panel/change_password_forced.html', contexto)
            
        if nueva_clave != confirmar_clave:
            contexto['error'] = 'Las contraseñas no coinciden.'
            return render(petición, 'admin_panel/change_password_forced.html', contexto)
            
        import re
        patron = r"^(?=.*[A-Z])(?=.*[0-9])(?=.*[@$!%*?&]).{8,}$"
        if not re.match(patron, nueva_clave):
            contexto['error'] = 'La contraseña debe tener al menos 8 caracteres, una mayúscula, un número y un carácter especial (@$!%*?&).'
            return render(petición, 'admin_panel/change_password_forced.html', contexto)
            
        # Actualizar contraseña en autenticación (Supabase / SQLite)
        from users.supabase_auth import actualizar_contraseña
        token = petición.session.get('token_acceso') or 'mock-local-token'
        
        datos, error = actualizar_contraseña(token, nueva_clave, user_id=perfil.id)
        if error:
            contexto['error'] = f'Error al cambiar la contraseña en el servidor: {error}'
            return render(petición, 'admin_panel/change_password_forced.html', contexto)
            
        # Actualizar perfil local
        from django.utils import timezone
        perfil.password_cambiada = True
        perfil.fecha_primer_login = timezone.now()
        perfil.save()
        
        # Actualizar estado de invitación
        if perfil.invitacion:
            inv = perfil.invitacion
            inv.estado = 'aceptada'
            inv.fecha_aceptacion = timezone.now()
            
            # Obtener IP de origen
            x_forwarded_for = petición.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip = x_forwarded_for.split(',')[0].strip()
            else:
                ip = petición.META.get('REMOTE_ADDR')
            inv.ip_origen = ip
            inv.save()
            
        messages.success(petición, '✅ Contraseña cambiada con éxito. Ya tienes acceso al Dashboard.')
        return redirect('/admin-panel/')
        
    return render(petición, 'admin_panel/change_password_forced.html', contexto)
